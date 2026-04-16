from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
from typing import Iterable
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


SUPPORTED_QUERYABLE_FILE_FORMATS = {"csv", "json", "parquet", "xml", "xlsx"}
DIRECT_QUERY_FILE_FORMATS = {"csv", "json", "parquet"}


@dataclass(frozen=True, slots=True)
class MaterializedQueryableFile:
    local_path: Path
    reader_format: str


def normalize_queryable_file_format(
    *,
    file_name: str,
    export_format: str = "",
    mime_type: str = "",
) -> str:
    normalized_export_format = str(export_format or "").strip().lower()
    if normalized_export_format in SUPPORTED_QUERYABLE_FILE_FORMATS:
        return normalized_export_format

    suffix = Path(str(file_name or "").strip()).suffix.lower().lstrip(".")
    if suffix in SUPPORTED_QUERYABLE_FILE_FORMATS:
        return suffix
    if suffix in {"jsonl", "ndjson"}:
        return "json"
    if suffix in {"xls", "xlsxm"}:
        return "xlsx"

    normalized_mime_type = str(mime_type or "").strip().lower()
    if "csv" in normalized_mime_type:
        return "csv"
    if "json" in normalized_mime_type:
        return "json"
    if "parquet" in normalized_mime_type:
        return "parquet"
    if "spreadsheetml" in normalized_mime_type or "excel" in normalized_mime_type:
        return "xlsx"
    if "xml" in normalized_mime_type:
        return "xml"

    raise ValueError(
        "Queryable file support currently handles CSV, JSON, Parquet, XML, and Excel files only."
    )


def materialize_queryable_file(
    *,
    root: Path,
    base_name: str,
    file_name: str,
    source_format: str,
    file_bytes: bytes,
) -> MaterializedQueryableFile:
    normalized_source_format = str(source_format or "").strip().lower()
    if normalized_source_format not in SUPPORTED_QUERYABLE_FILE_FORMATS:
        raise ValueError(f"Unsupported queryable file format: {source_format}")
    if not file_bytes:
        raise ValueError("The file is empty and cannot be queried.")

    root.mkdir(parents=True, exist_ok=True)
    if normalized_source_format in DIRECT_QUERY_FILE_FORMATS:
        suffix = _output_suffix(file_name=file_name, source_format=normalized_source_format)
        local_path = root / f"{base_name}{suffix}"
        local_path.write_bytes(file_bytes)
        return MaterializedQueryableFile(local_path=local_path, reader_format=normalized_source_format)

    local_path = root / f"{base_name}.csv"
    if normalized_source_format == "xml":
        columns, rows = _xml_columns_and_rows(file_bytes)
    elif normalized_source_format == "xlsx":
        columns, rows = _xlsx_columns_and_rows(file_bytes)
    else:
        raise ValueError(f"Unsupported queryable file format: {source_format}")

    _write_csv(local_path, columns=columns, rows=rows)
    return MaterializedQueryableFile(local_path=local_path, reader_format="csv")


def cached_queryable_output_path(
    *,
    root: Path,
    base_name: str,
    file_name: str,
    source_format: str,
) -> MaterializedQueryableFile:
    normalized_source_format = str(source_format or "").strip().lower()
    if normalized_source_format in DIRECT_QUERY_FILE_FORMATS:
        suffix = _output_suffix(file_name=file_name, source_format=normalized_source_format)
        return MaterializedQueryableFile(local_path=root / f"{base_name}{suffix}", reader_format=normalized_source_format)
    if normalized_source_format in {"xml", "xlsx"}:
        return MaterializedQueryableFile(local_path=root / f"{base_name}.csv", reader_format="csv")
    raise ValueError(f"Unsupported queryable file format: {source_format}")


def _output_suffix(*, file_name: str, source_format: str) -> str:
    existing_suffix = Path(str(file_name or "").strip()).suffix.lower()
    if existing_suffix:
        if source_format == "json" and existing_suffix in {".json", ".jsonl", ".ndjson"}:
            return existing_suffix
        if source_format == "csv" and existing_suffix in {".csv", ".tsv"}:
            return existing_suffix
        if source_format == "parquet" and existing_suffix == ".parquet":
            return existing_suffix
    return {
        "csv": ".csv",
        "json": ".json",
        "parquet": ".parquet",
    }[source_format]


def _write_csv(local_path: Path, *, columns: list[str], rows: Iterable[list[str]]) -> None:
    with local_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(row)


def _normalize_column_name(value: object, *, index: int, seen: set[str]) -> str:
    candidate = str(value or "").strip()
    if not candidate:
        candidate = f"column_{index + 1}"
    normalized = re.sub(r"\s+", " ", candidate).strip() or f"column_{index + 1}"
    unique = normalized
    next_index = 2
    while unique.lower() in seen:
        unique = f"{normalized}_{next_index}"
        next_index += 1
    seen.add(unique.lower())
    return unique


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return str(value)


def _xlsx_columns_and_rows(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)

        header_values = None
        for row in rows:
            if any(cell not in (None, "") for cell in row):
                header_values = list(row)
                break
        if header_values is None:
            raise ValueError("The Excel file is empty and cannot be queried.")

        seen_headers: set[str] = set()
        columns = [
            _normalize_column_name(value, index=index, seen=seen_headers)
            for index, value in enumerate(header_values)
        ]

        materialized_rows: list[list[str]] = []
        for row in rows:
            row_values = list(row)
            if not any(cell not in (None, "") for cell in row_values):
                continue
            if len(row_values) < len(columns):
                row_values.extend([None] * (len(columns) - len(row_values)))
            materialized_rows.append([_stringify_cell(value) for value in row_values[: len(columns)]])
        return columns, materialized_rows
    finally:
        workbook.close()


def _xml_columns_and_rows(file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        root = ET.fromstring(file_bytes)
    except ET.ParseError as exc:
        raise ValueError(f"The XML file is not well-formed and cannot be queried: {exc}") from exc

    row_elements = [element for element in list(root) if isinstance(element.tag, str)]
    if not row_elements:
        raise ValueError(
            "The XML file does not contain repeated row elements beneath the root and cannot be queried."
        )

    columns: list[str] = []
    seen_columns: set[str] = set()
    parsed_rows: list[dict[str, str]] = []
    for row_element in row_elements:
        row_values: dict[str, str] = {}
        child_elements = [element for element in list(row_element) if isinstance(element.tag, str)]
        if not child_elements:
            continue
        for child_element in child_elements:
            column_name = _strip_xml_namespace(str(child_element.tag or "").strip()) or "value"
            if column_name.lower() not in seen_columns:
                seen_columns.add(column_name.lower())
                columns.append(column_name)
            row_values[column_name] = _stringify_cell("".join(child_element.itertext()).strip())
        parsed_rows.append(row_values)

    if not columns:
        raise ValueError(
            "The XML file does not expose simple row/field elements and cannot be queried."
        )

    rows = [[row.get(column_name, "") for column_name in columns] for row in parsed_rows]
    return columns, rows


def _strip_xml_namespace(tag_name: str) -> str:
    normalized = str(tag_name or "").strip()
    if normalized.startswith("{") and "}" in normalized:
        return normalized.split("}", 1)[1]
    return normalized
