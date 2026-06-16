from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass
import contextlib
import io
from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import zipfile
from uuid import uuid4

import boto3
from botocore.exceptions import ClientError
import duckdb
from openpyxl import Workbook, load_workbook
import psycopg
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright


@dataclass(frozen=True, slots=True)
class FileSmokeSpec:
    ingestor_id: str
    label: str
    extension: str
    mime_type: str
    s3_prefix_root: str


def parse_args(description: str, s3_prefix_root: str) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--base-url", default="http://127.0.0.1:8000")
    parser.add_argument("--s3-endpoint", default="http://127.0.0.1:9000")
    parser.add_argument("--s3-access-key", default="minioadmin")
    parser.add_argument("--s3-secret-key", default="minioadmin")
    parser.add_argument("--bucket", default="vat-smoke-test")
    parser.add_argument(
        "--pg-dsn",
        default="host=127.0.0.1 port=5432 dbname=evo1_oltp user=evo1 password=evo1",
    )
    parser.add_argument("--pg-schema", default="stage")
    parser.add_argument("--s3-smoke-prefix-root", default=s3_prefix_root)
    parser.add_argument("--headless", action="store_true", default=True)
    parser.add_argument("--headed", dest="headless", action="store_false")
    parser.add_argument("--timeout-ms", type=int, default=60000)
    return parser.parse_args()


def zip_bytes(entries: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_STORED) as archive:
        for name, payload in entries.items():
            archive.writestr(name, payload)
    return buffer.getvalue()


def s3_client(args: argparse.Namespace):
    return boto3.client(
        "s3",
        endpoint_url=args.s3_endpoint,
        aws_access_key_id=args.s3_access_key,
        aws_secret_access_key=args.s3_secret_key,
    )


def ensure_bucket(client, bucket: str) -> None:
    try:
        client.head_bucket(Bucket=bucket)
    except ClientError as exc:
        status = int(exc.response.get("ResponseMetadata", {}).get("HTTPStatusCode", 0))
        code = str(exc.response.get("Error", {}).get("Code", ""))
        if status not in {400, 404} and code not in {"404", "NoSuchBucket"}:
            raise
        client.create_bucket(Bucket=bucket)


def delete_s3_prefix(client, bucket: str, prefix: str) -> None:
    normalized_prefix = "/".join(segment for segment in prefix.split("/") if segment)
    if not normalized_prefix:
        return
    list_prefix = f"{normalized_prefix}/"
    while True:
        response = client.list_objects_v2(Bucket=bucket, Prefix=list_prefix)
        objects = [{"Key": item["Key"]} for item in response.get("Contents", [])]
        if objects:
            client.delete_objects(Bucket=bucket, Delete={"Objects": objects})
        if not response.get("IsTruncated"):
            return


def payload_for(spec: FileSmokeSpec, row_id: int, name: str) -> bytes:
    if spec.ingestor_id == "json":
        return f'{{"id":{row_id},"name":"{name}"}}\n'.encode("utf-8")
    if spec.ingestor_id == "xml":
        return f"<rows><row><id>{row_id}</id><name>{name}</name></row></rows>".encode("utf-8")
    if spec.ingestor_id == "xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["id", "name"])
        worksheet.append([row_id, name])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()
    if spec.ingestor_id == "parquet":
        with TemporaryDirectory() as temp_dir:
            parquet_path = Path(temp_dir) / f"{name}.parquet"
            connection = duckdb.connect(":memory:")
            try:
                connection.execute(
                    "CREATE TABLE payload AS SELECT ?::INTEGER AS id, ?::VARCHAR AS name",
                    [row_id, name],
                )
                connection.execute(
                    f"COPY payload TO '{parquet_path.as_posix()}' (FORMAT PARQUET)"
                )
            finally:
                connection.close()
            return parquet_path.read_bytes()
    raise ValueError(f"Unsupported smoke spec: {spec.ingestor_id}")


def verify_s3_objects(
    client,
    bucket: str,
    spec: FileSmokeSpec,
    expected_payloads: dict[str, tuple[bytes, int, str]],
) -> None:
    with TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        for key, (expected_bytes, expected_id, expected_name) in expected_payloads.items():
            body = client.get_object(Bucket=bucket, Key=key)["Body"].read()
            if spec.ingestor_id == "json":
                if body != expected_bytes:
                    raise RuntimeError(f"Unexpected JSON S3 payload for {key!r}: {body!r}")
            elif spec.ingestor_id == "xml":
                if body != expected_bytes:
                    raise RuntimeError(f"Unexpected XML S3 payload for {key!r}: {body!r}")
            elif spec.ingestor_id == "xlsx":
                workbook = load_workbook(filename=io.BytesIO(body), read_only=True, data_only=True)
                try:
                    rows = list(workbook.active.iter_rows(values_only=True))
                finally:
                    workbook.close()
                if rows[:2] != [("id", "name"), (expected_id, expected_name)]:
                    raise RuntimeError(f"Unexpected XLSX S3 rows for {key!r}: {rows!r}")
            elif spec.ingestor_id == "parquet":
                parquet_path = temp_root / key.replace("/", "_")
                parquet_path.write_bytes(body)
                connection = duckdb.connect(":memory:")
                try:
                    rows = connection.execute(
                        f"SELECT id, name FROM read_parquet('{parquet_path.as_posix()}')"
                    ).fetchall()
                finally:
                    connection.close()
                if rows != [(expected_id, expected_name)]:
                    raise RuntimeError(f"Unexpected Parquet S3 rows for {key!r}: {rows!r}")


def verify_pg_tables(pg_dsn: str, schema: str, table_payloads: dict[str, tuple[int, str]]) -> None:
    with psycopg.connect(pg_dsn) as connection:
        with connection.cursor() as cursor:
            for table, (expected_id, expected_name) in table_payloads.items():
                cursor.execute(
                    """
                    select count(*)
                    from information_schema.tables
                    where table_schema = %s and table_name = %s
                    """,
                    (schema, table),
                )
                if cursor.fetchone()[0] != 1:
                    raise RuntimeError(f"PostgreSQL import did not create {schema}.{table}.")
                cursor.execute(f'select id, name from "{schema}"."{table}" order by id')
                rows = cursor.fetchall()
                if rows != [(expected_id, expected_name)]:
                    raise RuntimeError(
                        f"Unexpected PostgreSQL rows for {schema}.{table}: {rows!r}"
                    )


def drop_pg_tables(pg_dsn: str, schema: str, tables: list[str]) -> None:
    if not tables:
        return
    with psycopg.connect(pg_dsn, autocommit=True) as connection:
        with connection.cursor() as cursor:
            for table in tables:
                cursor.execute(f'drop table if exists "{schema}"."{table}"')


async def open_ingestion_landing(page, base_url: str, timeout_ms: int) -> None:
    await page.goto(
        f"{base_url.rstrip('/')}/ingestion-workbench",
        wait_until="domcontentloaded",
        timeout=timeout_ms,
    )
    await page.locator("[data-ingestion-workbench-page]").wait_for(
        state="visible",
        timeout=timeout_ms,
    )
    await page.wait_for_timeout(2000)


async def assert_landing_search(page, args: argparse.Namespace) -> None:
    await open_ingestion_landing(page, args.base_url, args.timeout_ms)
    search = page.locator("[data-ingestion-search-input]")
    await search.fill("json")
    await page.locator('[data-ingestion-tile="json"]').wait_for(
        state="visible",
        timeout=args.timeout_ms,
    )
    try:
        await page.locator('[data-ingestion-tile="parquet"]').wait_for(
            state="hidden",
            timeout=3000,
        )
    except PlaywrightTimeoutError:
        raise RuntimeError("Searching json did not hide unrelated Parquet ingestor.")
    await search.fill("xml")
    xml_tile = page.locator('[data-ingestion-tile="xml"]')
    await xml_tile.wait_for(state="visible", timeout=args.timeout_ms)
    try:
        await page.locator('[data-ingestion-tile="json"]').wait_for(
            state="hidden",
            timeout=3000,
        )
    except PlaywrightTimeoutError:
        raise RuntimeError("Searching xml did not hide unrelated JSON ingestor.")
    await xml_tile.click()
    xml_tooltip = page.locator('[data-file-tooltip="xml"]')
    await xml_tooltip.wait_for(state="visible", timeout=args.timeout_ms)
    tooltip_text = await xml_tooltip.get_attribute("title")
    expected = "XML support expects a simple table-like document"
    if expected not in str(tooltip_text or ""):
        raise RuntimeError(f"XML tooltip missing limitation copy: {tooltip_text!r}")
    await page.locator('#xml-ingestion-panel [data-close-ingestion-entry]').click()
    await search.fill("")


async def open_ingestor(page, args: argparse.Namespace, spec: FileSmokeSpec) -> None:
    await open_ingestion_landing(page, args.base_url, args.timeout_ms)
    tile = page.locator(f'[data-ingestion-tile="{spec.ingestor_id}"]').first
    form = page.locator(
        f'[data-file-ingestion-form][data-file-ingestor-id="{spec.ingestor_id}"]'
    )
    for _attempt in range(5):
        await tile.click()
        try:
            await form.wait_for(state="visible", timeout=2000)
            break
        except PlaywrightTimeoutError:
            await page.wait_for_timeout(500)
    await form.wait_for(state="visible", timeout=args.timeout_ms)


async def assert_success_dialog(
    page,
    timeout_ms: int,
    expected_fragments: list[str],
) -> None:
    dialog = page.locator("[data-message-dialog]")
    await dialog.wait_for(state="visible", timeout=timeout_ms)
    dialog_text = (await dialog.text_content() or "").strip()
    missing = [fragment for fragment in expected_fragments if fragment not in dialog_text]
    if missing:
        raise RuntimeError(f"Success dialog is missing {missing!r}: {dialog_text!r}")
    await page.locator("[data-message-submit]").click()
    await dialog.wait_for(state="hidden", timeout=timeout_ms)


async def poll_upload_session_result(
    page,
    base_url: str,
    spec: FileSmokeSpec,
    session_id: str,
    timeout_ms: int,
) -> dict[str, object]:
    deadline = asyncio.get_running_loop().time() + (timeout_ms / 1000)
    session_url = (
        f"{base_url.rstrip('/')}/api/ingestion/{spec.ingestor_id}/upload-sessions/{session_id}"
    )
    while asyncio.get_running_loop().time() < deadline:
        response = await page.context.request.get(
            session_url,
            headers={"Accept": "application/json"},
        )
        if not response.ok:
            raise RuntimeError(f"Upload session polling failed with HTTP {response.status}.")
        payload = await response.json()
        result = payload.get("result") if isinstance(payload, dict) else None
        if isinstance(result, dict):
            return result
        if isinstance(payload, dict) and payload.get("status") == "failed":
            raise RuntimeError(f"Upload session failed: {payload!r}")
        await page.wait_for_timeout(500)
    raise RuntimeError("Timed out waiting for upload session import result.")


async def expect_complete_payload(page, args: argparse.Namespace, spec: FileSmokeSpec):
    async with page.expect_response(
        lambda response: response.request.method == "POST"
        and f"/api/ingestion/{spec.ingestor_id}/upload-sessions/" in response.url
        and response.url.endswith("/complete"),
        timeout=args.timeout_ms,
    ) as response_info:
        await page.locator(
            f'[data-file-ingestion-form][data-file-ingestor-id="{spec.ingestor_id}"] [data-file-import-submit]'
        ).click()
    response = await response_info.value
    if not response.ok:
        raise RuntimeError(f"Completion request failed with HTTP {response.status}.")
    payload = await response.json()
    if isinstance(payload, dict) and isinstance(payload.get("result"), dict):
        return payload["result"]
    if isinstance(payload, dict) and "targetId" in payload:
        return payload
    session_id = str(payload.get("sessionId") or "") if isinstance(payload, dict) else ""
    if session_id:
        return await poll_upload_session_result(
            page,
            args.base_url,
            spec,
            session_id,
            args.timeout_ms,
        )
    raise RuntimeError(f"Unexpected completion response payload: {payload!r}")


async def set_target_and_file(
    page,
    args: argparse.Namespace,
    spec: FileSmokeSpec,
    target_id: str,
    prefix_or_table_prefix: str,
    payload_name: str,
    payload: bytes,
) -> None:
    await open_ingestor(page, args, spec)
    form = page.locator(
        f'[data-file-ingestion-form][data-file-ingestor-id="{spec.ingestor_id}"]'
    )
    await form.locator(f'[data-file-target-option][value="{target_id}"]').check()
    if target_id == "s3":
        await form.locator('[data-file-config-panel="s3"] [data-file-s3-bucket]').fill(
            args.bucket
        )
        await form.locator('[data-file-config-panel="s3"] [data-file-s3-prefix]').fill(
            prefix_or_table_prefix
        )
    else:
        await form.locator(f'[data-file-config-panel="{target_id}"] [data-file-schema-name]').fill(
            args.pg_schema
        )
        await form.locator(f'[data-file-config-panel="{target_id}"] [data-file-table-prefix]').fill(
            prefix_or_table_prefix
        )
    await form.locator("[data-file-input]").set_input_files(
        files=[{"name": payload_name, "mimeType": spec.mime_type, "buffer": payload}]
    )


async def import_to_s3(
    page,
    args: argparse.Namespace,
    spec: FileSmokeSpec,
    unique_id: str,
    *,
    archive: bool,
) -> tuple[str, dict[str, tuple[bytes, int, str]]]:
    mode = "zip" if archive else "direct"
    prefix = f"{args.s3_smoke_prefix_root.strip('/')}/{unique_id}/{mode}"
    if archive:
        entries = {
            f"beta{spec.extension}": (payload_for(spec, 2, "beta"), 2, "beta"),
            f"nested/gamma{spec.extension}": (payload_for(spec, 3, "gamma"), 3, "gamma"),
        }
        payload = zip_bytes({name: item[0] for name, item in entries.items()})
        payload_name = f"playwright-{spec.ingestor_id}-batch.zip"
        expected = {
            f"{prefix}/{Path(name).name}": (item[0], item[1], item[2])
            for name, item in entries.items()
        }
    else:
        payload_name = f"alpha{spec.extension}"
        payload_bytes = payload_for(spec, 1, "alpha")
        payload = payload_bytes
        expected = {f"{prefix}/{payload_name}": (payload_bytes, 1, "alpha")}

    await set_target_and_file(
        page,
        args,
        spec,
        "s3",
        prefix,
        payload_name,
        payload,
    )
    if archive:
        preview = page.locator(
            f'[data-file-ingestion-form][data-file-ingestor-id="{spec.ingestor_id}"] '
            "[data-file-preview-root] .ingestion-csv-preview-card"
        ).first
        preview_text = ""
        for _attempt in range(20):
            preview_text = (await preview.text_content() or "")
            if f"2 {spec.label} file(s)" in preview_text:
                break
            await page.wait_for_timeout(250)
        if f"2 {spec.label} file(s)" not in preview_text:
            raise RuntimeError(f"ZIP preview did not report two files: {preview_text!r}")

    result = await expect_complete_payload(page, args, spec)
    if result.get("targetId") != "s3":
        raise RuntimeError(f"Unexpected S3 target: {result!r}")
    if result.get("importedCount") != len(expected) or result.get("failedCount") != 0:
        raise RuntimeError(f"Unexpected S3 import counts: {result!r}")
    result_list = page.locator(
        f'[data-file-ingestion-form][data-file-ingestor-id="{spec.ingestor_id}"] [data-file-result-list]'
    )
    await result_list.locator(".ingestion-csv-result-card-imported").nth(len(expected) - 1).wait_for(
        state="visible",
        timeout=args.timeout_ms,
    )
    result_text = (await result_list.text_content() or "").strip()
    missing = [fragment for fragment in expected if fragment not in result_text]
    if missing:
        raise RuntimeError(f"S3 result UI is missing {missing!r}: {result_text!r}")
    await assert_success_dialog(
        page,
        args.timeout_ms,
        [f"{spec.label} import finished", f"{len(expected)} file(s) processed", "S3 Object Storage"],
    )
    return prefix, expected


async def import_to_pg(
    page,
    args: argparse.Namespace,
    spec: FileSmokeSpec,
    unique_id: str,
    *,
    archive: bool,
) -> dict[str, tuple[int, str]]:
    mode = "zip" if archive else "direct"
    table_prefix = f"pw{spec.ingestor_id}_{unique_id}_{mode}"
    if archive:
        payload_name = f"playwright-{spec.ingestor_id}-batch.zip"
        payload = zip_bytes(
            {
                f"beta{spec.extension}": payload_for(spec, 2, "beta"),
                f"nested/gamma{spec.extension}": payload_for(spec, 3, "gamma"),
            }
        )
        expected_tables = {
            f"{table_prefix}_beta": (2, "beta"),
            f"{table_prefix}_gamma": (3, "gamma"),
        }
    else:
        payload_name = f"alpha{spec.extension}"
        payload = payload_for(spec, 1, "alpha")
        expected_tables = {f"{table_prefix}_alpha": (1, "alpha")}

    await set_target_and_file(
        page,
        args,
        spec,
        "pg_oltp",
        table_prefix,
        payload_name,
        payload,
    )
    result = await expect_complete_payload(page, args, spec)
    if result.get("targetId") != "pg_oltp":
        raise RuntimeError(f"Unexpected PostgreSQL target: {result!r}")
    if result.get("importedCount") != len(expected_tables) or result.get("failedCount") != 0:
        raise RuntimeError(f"Unexpected PostgreSQL import counts: {result!r}")

    expected_relations = [f"{args.pg_schema}.{table}" for table in expected_tables]
    result_list = page.locator(
        f'[data-file-ingestion-form][data-file-ingestor-id="{spec.ingestor_id}"] [data-file-result-list]'
    )
    await result_list.locator(".ingestion-csv-result-card-imported").nth(
        len(expected_tables) - 1
    ).wait_for(state="visible", timeout=args.timeout_ms)
    result_text = (await result_list.text_content() or "").strip()
    missing = [fragment for fragment in expected_relations if fragment not in result_text]
    if missing:
        raise RuntimeError(f"PostgreSQL result UI is missing {missing!r}: {result_text!r}")
    await assert_success_dialog(
        page,
        args.timeout_ms,
        [f"{spec.label} import finished", f"{len(expected_tables)} file(s) processed", "PostgreSQL OLTP"],
    )
    return expected_tables


async def run_smoke(args: argparse.Namespace, spec: FileSmokeSpec) -> int:
    unique_id = uuid4().hex[:8]
    client = s3_client(args)
    ensure_bucket(client, args.bucket)
    delete_s3_prefix(client, args.bucket, args.s3_smoke_prefix_root)
    s3_prefixes: list[str] = []
    pg_tables: list[str] = []
    console_messages: list[str] = []
    responses: list[tuple[str, str, int]] = []

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=args.headless)
        page = await browser.new_page(
            viewport={"width": 1440, "height": 1200},
            base_url=args.base_url.rstrip("/"),
        )
        page.on("console", lambda msg: console_messages.append(f"console:{msg.type}:{msg.text}"))
        page.on("pageerror", lambda exc: console_messages.append(f"pageerror:{exc}"))
        page.on(
            "response",
            lambda resp: responses.append((resp.request.method, resp.url, resp.status)),
        )
        try:
            await assert_landing_search(page, args)
            for archive in (False, True):
                s3_prefix, expected = await import_to_s3(
                    page,
                    args,
                    spec,
                    unique_id,
                    archive=archive,
                )
                s3_prefixes.append(s3_prefix)
                verify_s3_objects(client, args.bucket, spec, expected)
            for archive in (False, True):
                expected_tables = await import_to_pg(
                    page,
                    args,
                    spec,
                    unique_id,
                    archive=archive,
                )
                pg_tables.extend(expected_tables)
                verify_pg_tables(args.pg_dsn, args.pg_schema, expected_tables)
        except (ClientError, PlaywrightTimeoutError, RuntimeError, psycopg.Error) as exc:
            print(str(exc), file=sys.stderr)
            for method, url, status in responses:
                if f"/api/ingestion/{spec.ingestor_id}/upload-sessions" in url:
                    print(f"HTTP {method} {status} {url}", file=sys.stderr)
            for message in console_messages:
                print(message, file=sys.stderr)
            await browser.close()
            return 1
        finally:
            await browser.close()
            for prefix in s3_prefixes:
                with contextlib.suppress(Exception):
                    delete_s3_prefix(client, args.bucket, prefix)
            with contextlib.suppress(Exception):
                drop_pg_tables(args.pg_dsn, args.pg_schema, pg_tables)

    print(f"Playwright {spec.label} ingestion real backend smoke passed for id {unique_id}.")
    return 0
